# -*- coding: utf-8 -*-
"""
Export SQL Server table structure to an Excel file.

PyCharm usage:
1. Open Settings > Project > Python Interpreter.
2. Install dependencies in the current interpreter:
   pip install openpyxl sqlalchemy pyodbc
3. Update the configuration section below.
4. Right-click this file and choose Run.
"""

from __future__ import annotations

import pathlib
import sys
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from sqlalchemy import create_engine, text
    from sqlalchemy.engine import URL
except ModuleNotFoundError as exc:
    missing_package = exc.name
    print(f"缺少 Python 包：{missing_package}")
    print("请在 PyCharm 当前项目解释器里安装依赖：")
    print("pip install openpyxl sqlalchemy pyodbc")
    raise SystemExit(1) from exc


try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass


# ==================== 配置区域 ====================
# 连接方式：
# - "windows": 使用 Windows 身份验证，适合本机 SQL Server / SSMS 里选择“Windows 身份验证”的情况
# - "sql": 使用 SQL Server 身份验证，适合使用 sa 或其他数据库用户名密码的情况
AUTH_MODE = "windows"

# SQL Server 地址。
# 常见写法：
# - 默认实例：localhost
# - SQLEXPRESS：localhost\SQLEXPRESS
# - 指定 IP：127.0.0.1
SERVER = r"localhost"

# 如果你的 SQL Server 使用固定端口，填 1433；如果 SERVER 写了 localhost\SQLEXPRESS，通常填 None。
PORT: int | None = 1433

DATABASE = "master"
SCHEMA = "dbo"

# 只有 AUTH_MODE = "sql" 时才会使用下面两个值。
# 注意：这里必须填 SQL Server 登录名，不要填电脑名。
USERNAME = "sa"
PASSWORD = "your_password"

# 需要导出的表名。留空会先列出当前数据库 dbo 下的全部表名。
TABLES: list[str] = []

ODBC_DRIVER = "ODBC Driver 17 for SQL Server"

SCRIPT_DIR = pathlib.Path(__file__).resolve().parent
OUTPUT_FILE = SCRIPT_DIR / "数据库pds.xlsx"

HEADERS = ["列名*", "列中文名*", "类型*", "可为空?", "主键*", "长度", "精度"]
# =================================================


TABLE_COMMENT_QUERY = text(
    """
    SELECT
        CAST(ep.value AS NVARCHAR(MAX)) AS table_comment
    FROM sys.tables t
    LEFT JOIN sys.extended_properties ep
        ON ep.major_id = t.object_id
        AND ep.minor_id = 0
        AND ep.name = 'MS_Description'
    WHERE t.name = :table_name
      AND SCHEMA_NAME(t.schema_id) = :schema
    """
)

COLUMNS_QUERY = text(
    """
    SELECT
        c.ordinal_position,
        c.column_name,
        ISNULL(CAST(ep.value AS NVARCHAR(MAX)), '') AS column_comment,
        c.data_type,
        CASE WHEN c.is_nullable = 'YES' THEN N'是' ELSE N'否' END AS is_nullable,
        CASE WHEN pk.column_name IS NOT NULL THEN N'是' ELSE N'否' END AS is_primary_key,
        c.character_maximum_length AS length_value,
        c.numeric_precision AS precision_value
    FROM information_schema.columns c
    LEFT JOIN (
        SELECT
            kcu.table_schema,
            kcu.table_name,
            kcu.column_name
        FROM information_schema.table_constraints tc
        JOIN information_schema.key_column_usage kcu
            ON tc.constraint_name = kcu.constraint_name
            AND tc.table_schema = kcu.table_schema
            AND tc.table_name = kcu.table_name
        WHERE tc.constraint_type = 'PRIMARY KEY'
    ) pk
        ON c.table_schema = pk.table_schema
        AND c.table_name = pk.table_name
        AND c.column_name = pk.column_name
    LEFT JOIN sys.columns sc
        ON OBJECT_ID(QUOTENAME(c.table_schema) + '.' + QUOTENAME(c.table_name)) = sc.object_id
        AND c.column_name = sc.name
    LEFT JOIN sys.extended_properties ep
        ON ep.major_id = sc.object_id
        AND ep.minor_id = sc.column_id
        AND ep.name = 'MS_Description'
    WHERE c.table_catalog = :database
      AND c.table_schema = :schema
      AND c.table_name = :table_name
    ORDER BY c.ordinal_position
    """
)

LIST_TABLES_QUERY = text(
    """
    SELECT TABLE_NAME
    FROM information_schema.tables
    WHERE TABLE_SCHEMA = :schema
      AND TABLE_TYPE = 'BASE TABLE'
    ORDER BY TABLE_NAME
    """
)


def safe_sheet_name(name: str, used_names: set[str]) -> str:
    sheet_name = name[:31]
    if sheet_name not in used_names:
        used_names.add(sheet_name)
        return sheet_name

    base = sheet_name[:28]
    index = 1
    while True:
        candidate = f"{base}_{index}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        index += 1


def write_required_label(ws: Any, cell: str, label: str) -> None:
    normal_font = Font(name="宋体", size=12, color="000000")
    red_font = Font(name="宋体", size=12, color="FF0000")
    ws[cell] = label
    ws[cell].font = red_font if label.endswith("*") else normal_font


def style_sheet(ws: Any) -> None:
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A5"

    widths = {
        "A": 22,
        "B": 26,
        "C": 18,
        "D": 14,
        "E": 14,
        "F": 12,
        "G": 12,
        "H": 2,
        "I": 30,
        "J": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in range(1, 5):
        ws.row_dimensions[row].height = 28

    for row in ws.iter_rows(min_row=1, max_row=max(ws.max_row, 8), min_col=1, max_col=10):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name="宋体", size=12)

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    for cell in ws[4][0:7]:
        cell.fill = header_fill
        cell.font = Font(name="宋体", size=12, bold=True)

    ws["I4"] = "说明：带 * 的列为必填内容"
    ws["I4"].alignment = Alignment(horizontal="left", vertical="center")
    ws["I4"].font = Font(name="宋体", size=12)


def to_allowed_data_type(data_type: str) -> str:
    normalized_type = data_type.lower()
    if normalized_type in {"char", "nchar", "varchar", "nvarchar", "uniqueidentifier"}:
        return "字符型"
    if normalized_type in {"text", "ntext", "xml", "json"}:
        return "文本型"
    if normalized_type in {"tinyint", "smallint", "int", "bigint", "bit"}:
        return "整数型"
    if normalized_type in {"real", "float"}:
        return "浮点型"
    if normalized_type in {"decimal", "numeric", "money", "smallmoney"}:
        return "数字型"
    if normalized_type == "date":
        return "日期型"
    if normalized_type in {
        "time",
        "datetime",
        "datetime2",
        "datetimeoffset",
        "smalldatetime",
        "timestamp",
        "rowversion",
    }:
        return "时间戳"
    if normalized_type in {"binary", "varbinary", "image"}:
        return "二进制"
    if normalized_type in {"geography", "geometry"}:
        return "地理坐标"
    return "文本型"
def write_table_sheet(
    wb: Workbook,
    table_name: str,
    table_comment: str | None,
    rows: list[dict[str, Any]],
    used_sheet_names: set[str],
) -> None:
    ws = wb.create_sheet(title=safe_sheet_name(table_name, used_sheet_names))

    write_required_label(ws, "A1", "表名*")
    ws["B1"] = table_name
    write_required_label(ws, "A2", "表注释")
    ws["B2"] = table_comment or ""

    for col_index, header in enumerate(HEADERS, start=1):
        ws.cell(row=4, column=col_index, value=header)

    for row_index, row in enumerate(rows, start=5):
        values = [
            row["column_name"],
            row["column_comment"],
            to_allowed_data_type(row["data_type"]),
            row["is_nullable"],
            row["is_primary_key"],
            row["length_value"],
            row["precision_value"],
        ]
        for col_index, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=col_index, value=value)

    style_sheet(ws)


def build_engine():
    query = {
        "driver": ODBC_DRIVER,
        "TrustServerCertificate": "yes",
    }

    if AUTH_MODE == "windows":
        query["Trusted_Connection"] = "yes"
        url = URL.create(
            "mssql+pyodbc",
            host=SERVER,
            port=PORT,
            database=DATABASE,
            query=query,
        )
    elif AUTH_MODE == "sql":
        url = URL.create(
            "mssql+pyodbc",
            username=USERNAME,
            password=PASSWORD,
            host=SERVER,
            port=PORT,
            database=DATABASE,
            query=query,
        )
    else:
        raise ValueError('AUTH_MODE 只能是 "windows" 或 "sql"')

    return create_engine(url, pool_pre_ping=True)


def list_all_tables(engine) -> list[str]:
    with engine.connect() as conn:
        result = conn.execute(LIST_TABLES_QUERY, {"schema": SCHEMA})
        return [row[0] for row in result]


def main() -> None:
    engine = build_engine()

    try:
        with engine.connect() as conn:
            version = conn.execute(text("SELECT @@VERSION AS version")).scalar()
            print("SQL Server 连接成功")
            print(f"版本：{version[:80]}...")

        if not TABLES:
            print(f"\n未指定表名，正在列出 {DATABASE}.{SCHEMA} 下的表：")
            all_tables = list_all_tables(engine)
            if not all_tables:
                print("没有找到任何表，请检查 DATABASE 和 SCHEMA 配置。")
                return

            for index, table_name in enumerate(all_tables, start=1):
                print(f"{index}. {table_name}")

            print("\n请把需要导出的表名填入 TABLES，例如：")
            print('TABLES = ["表名1", "表名2"]')
            return

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        used_sheet_names: set[str] = set()
        exported_count = 0

        with engine.connect() as conn:
            for table in TABLES:
                if not table.strip():
                    continue

                try:
                    table_comment = conn.execute(
                        TABLE_COMMENT_QUERY,
                        {"schema": SCHEMA, "table_name": table},
                    ).scalar()

                    result = conn.execute(
                        COLUMNS_QUERY,
                        {
                            "database": DATABASE,
                            "schema": SCHEMA,
                            "table_name": table,
                        },
                    )
                    rows = [dict(row) for row in result.mappings()]

                    if not rows:
                        print(f"跳过：没有查到表结构，请检查表名或 schema：{SCHEMA}.{table}")
                        continue

                    write_table_sheet(wb, table, table_comment, rows, used_sheet_names)
                    exported_count += 1
                    print(f"已导出：{table}，共 {len(rows)} 列")
                except Exception as exc:
                    print(f"导出失败：{table}，错误：{exc}")

        if exported_count == 0:
            print("\n没有导出任何表，请检查 TABLES、DATABASE、SCHEMA 是否正确。")
            return

        wb.save(OUTPUT_FILE)
        print(f"\n导出完成：{OUTPUT_FILE}")
        print(f"成功导出 {exported_count} 个表。")
    except Exception as exc:
        print(f"\n错误：{exc}")
        print("\n请重点检查：")
        print("1. PyCharm 使用的解释器是否已安装：openpyxl sqlalchemy pyodbc")
        print(f"2. 是否已安装 Microsoft {ODBC_DRIVER}")
        print("3. SERVER、PORT、DATABASE、SCHEMA 是否正确")
        print("4. 如果用 Windows 身份验证，AUTH_MODE 应为 windows")
        print("5. 如果用 SQL Server 身份验证，AUTH_MODE 应为 sql，USERNAME/PASSWORD 要填 SQL 登录名和密码")
    finally:
        engine.dispose()


if __name__ == "__main__":
    main()

